import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import re
import os

def run_step3_pole_sync(file_path, parsed_data=None):
    """
    Tahap 3: Sinkronisasi Kolom A-K.
    Memproses Kapasitas, Jalur (Line), Tube, Core, Data Tiang (Pole), dan Grup G.
    """
    
    # --- 1. KONFIGURASI WARNA TUBE (ANSI Standard) ---
    tube_styles = {
        1: ("Blue", "0000FF", "FFFFFF"),
        2: ("Orange", "FF8C00", "000000"),
        3: ("Green", "008000", "FFFFFF"),
        4: ("Brown", "A52A2A", "FFFFFF")
    }

    # Mencari file hasil Tahap 1 di root folder (sejajar launcher.py)
    source_file = "1.Hasil_Convert.xlsx"
    if not os.path.exists(source_file):
        raise FileNotFoundError(f"❌ Sumber data {source_file} tidak ditemukan. Pastikan Tahap 1 sudah selesai.")

    # --- 2. PRE-PROCESS DATA DARI SHEET CABLE (CAPACITY & LINE) ---
    cable_list = []
    line_list = []
    try:
        df_cable = pd.read_excel(source_file, sheet_name='CABLE')
        for _, row in df_cable.iterrows():
            name_str = str(row['Name'])
            cap_match = re.search(r'(\d+C/\d+T)', name_str)
            line_match = re.search(r'(LINE\s+[A-Z])', name_str.upper())

            cable_list.append(cap_match.group(1) if cap_match else "")
            line_list.append(line_match.group(1) if line_match else "")
            
    except Exception as e:
        print(f"WARNING: Gagal memproses data Capacity/Line: {e}")

    # --- 3. AMBIL DATA POLE DARI HASIL TAHAP 1 ---
    try:
        df_source = pd.read_excel(source_file, sheet_name='FAT & POLE')
        pole_data = []
        for _, row in df_source.iterrows():
            if pd.notna(row.get('POLE Name')):
                pole_data.append({
                    'name': row['POLE Name'],
                    'lat': row['Latitude'],
                    'long': row['Longitude']
                })
    except Exception as e:
        raise Exception(f"Gagal membaca sumber data tiang: {e}")

    # --- 4. LOAD HPDB UNTUK PROSES INJEKSI FINAL ---
    wb = load_workbook(file_path)
    ws = wb["HOMEPASS DATABASE"]
    
    # Deteksi baris terakhir yang valid berdasarkan kolom AV (FAT Code)
    df_hpdb = pd.read_excel(file_path, sheet_name='HOMEPASS DATABASE', header=None)
    av_col = df_hpdb.iloc[:, 47] # Kolom AV
    real_max_row = av_col.last_valid_index() + 1 if av_col.notna().any() else 10

    # Variabel Kontrol Logika
    p_idx = -1
    c_idx = -1
    current_pole = None
    previous_av = None
    current_prefix = None
    fdt_port_counter = 0 
    core_counter = 1
    cores_in_tube = 0
    MAX_CORES_PER_TUBE = 10 

    g_number = 1
    g_occurrence = 0

    # --- 5. LOOPING INJEKSI DATA (BARIS 10 KE BAWAH) ---
    for r in range(10, real_max_row + 1):
        av_value = ws.cell(row=r, column=48).value  # Kolom AV
        h_value = ws.cell(row=r, column=8).value    # Kolom H

        if av_value is None or str(av_value).strip() == "":
            continue

        # A. LOGIKA PREFIX & SYNC TIANG
        prefix = str(av_value)[0].upper()
        
        # Reset counter port jika terdeteksi FAT awal (A01)
        if str(av_value).strip().upper() == "A01" and previous_av != "A01":
            fdt_port_counter = 0

        # Reset Core jika Prefix kabel berganti (A, B, C...)
        if prefix != current_prefix:
            core_counter = 1
            cores_in_tube = 0
            current_prefix = prefix
            c_idx += 1 

        # Update data tiang jika kode FAT berganti
        if av_value != previous_av:
            p_idx += 1
            previous_av = av_value
            current_pole = pole_data[p_idx] if p_idx < len(pole_data) else None

        # Injeksi Data Tiang ke Kolom I, J, K
        if current_pole:
            ws.cell(row=r, column=9).value = current_pole['name']
            ws.cell(row=r, column=10).value = current_pole['lat']
            ws.cell(row=r, column=11).value = current_pole['long']
        else:
            ws.cell(row=r, column=9).value = "N/A"

        # B. LOGIKA PENOMORAN PORT (B) & GRUP G (A)
        if h_value in [1, 2]:
            fdt_port_counter += 1
            ws.cell(row=r, column=2).value = fdt_port_counter

            if fdt_port_counter == 1:
                g_number = 1
                g_occurrence = 0
            
            # Isi Kolom A (Grup G) hanya pada baris ganjil
            if fdt_port_counter % 2 != 0:
                ws.cell(row=r, column=1).value = f"G{g_number}"
                g_occurrence += 1
                
                # Ganti grup setiap 16 port (8 ganjil)
                if g_occurrence >= 8:
                    g_number += 1
                    g_occurrence = 0
            else:
                ws.cell(row=r, column=1).value = None

            # C. LOGIKA TUBE & CORE
            tube_num = ((core_counter - 1) // 12) + 1
            
            # Injeksi Capacity & Line pada port pertama tiap kabel
            if h_value == 1:
                if 0 <= c_idx < len(cable_list):
                    ws.cell(row=r, column=3).value = line_list[c_idx]
                    ws.cell(row=r, column=4).value = cable_list[c_idx]

            ws.cell(row=r, column=6).value = core_counter
            
            # Visual Styling Tube (Kolom E)
            if tube_num in tube_styles:
                _, bg_hex, font_hex = tube_styles[tube_num]
                cell_e = ws.cell(row=r, column=5)
                cell_e.value = tube_num 
                cell_e.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
                cell_e.font = Font(color=font_hex, bold=True)

            cores_in_tube += 1
            core_counter += 1

            # Logika pembatasan core per tube (Max 10, skip 2 core cadangan)
            if cores_in_tube == MAX_CORES_PER_TUBE:
                core_counter += 2 
                cores_in_tube = 0

    # --- 6. SIMPAN HASIL FINAL ---
    wb.save(file_path)
    return file_path