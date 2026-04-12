import zipfile
import os
import math
import shutil
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font

# Konfigurasi Internal Engine
STANDARD_HEADERS = ['Folder Name', 'Name', 'Latitude', 'Longitude', 'Description', 'Total Route (M)']
FP_HEADERS = ['FAT Name', 'POLE Name', 'Latitude', 'Longitude', 'Description', 'Total Route (M)']
FUZZY_MAPPING = {"CABLE": ["CABLE", "DISTRIBUTION"], "HP COVER": ["HP COVER"]}

def haversine(lat1, lon1, lat2, lon2):
    try:
        lat1, lon1, lat2, lon2 = map(float, [lat1, lon1, lat2, lon2])
        R = 6371000 
        phi1, phi2 = math.radians(lat1), math.radians(lat2)
        dphi, dlambda = math.radians(lat2-lat1), math.radians(lon2-lon1)
        a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlambda/2)**2
        return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    except: return float('inf')

def apply_styling(ws, df):
    ws.sheet_view.zoomScale = 70
    f = Font(name='Calibri', size=12)
    for i in range(len(df.columns)):
        ws.column_dimensions[chr(65 + i)].width = 22
    for row in ws.iter_rows(min_row=1, max_row=len(df)+1):
        for cell in row: cell.font = f

def run_conversion(kmz_path, template_path):
    """
    REVISI: Menangani file .bin dari backend dengan menyalinnya 
    ke ekstensi .xlsx terlebih dahulu agar didukung openpyxl.
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"❌ File template backend tidak ditemukan di: {template_path}")

    # 1. Tentukan nama file output di root
    output_file = "1.Hasil_Convert.xlsx"
    
    # 2. COPY TERLEBIH DAHULU (PENTING!)
    # Kita ubah .bin menjadi .xlsx saat proses copy agar openpyxl mau membaca
    shutil.copyfile(template_path, output_file)

    # 3. Ekstraksi KML dari KMZ
    with zipfile.ZipFile(kmz_path, 'r') as z:
        kml_name = next(n for n in z.namelist() if n.endswith('.kml'))
        content = z.read(kml_name).decode('utf-8', errors='ignore')
    
    # 4. Ambil daftar sheet dari file yang sudah di-copy (format .xlsx)
    wb_t = load_workbook(output_file, read_only=True)
    temp_names = [n.strip() for n in wb_t.sheetnames]
    wb_t.close()

    soup = BeautifulSoup(content, 'lxml-xml')
    data_by_sheet = {name: [] for name in temp_names}
    
    for placemark in soup.find_all('Placemark'):
        target_sheet, folder_val = None, ""
        curr = placemark
        while curr is not None and target_sheet is None:
            if curr.name == 'Folder':
                fn = curr.find('name', recursive=False).text.strip() if curr.find('name', recursive=False) else ""
                if not folder_val: folder_val = fn
                for k, v in FUZZY_MAPPING.items():
                    if any(kw.upper() in fn.upper() for kw in v): target_sheet = k
                if not target_sheet and fn in temp_names: target_sheet = fn
            curr = curr.parent

        if target_sheet:
            coords = placemark.find('Point').find('coordinates').text.strip().split(',') if placemark.find('Point') else ["",""]
            data_by_sheet[target_sheet].append({
                'Folder Name': folder_val if target_sheet == "HP COVER" else "",
                'Name': placemark.find('name').text if placemark.find('name') else "",
                'Latitude': coords[1] if len(coords) > 1 else "",
                'Longitude': coords[0] if len(coords) > 0 else "",
                'Description': placemark.find('description').text if placemark.find('description') else "",
                'Total Route (M)': ''
            })

    parsed_dfs = {k: pd.DataFrame(v) for k, v in data_by_sheet.items() if v}
    
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        for s_name, df in parsed_dfs.items():
            available_headers = [h for h in STANDARD_HEADERS if h in df.columns]
            df[available_headers].to_excel(writer, sheet_name=s_name, index=False)
            
            if s_name == "CABLE":
                for r in range(2, len(df)+2):
                    writer.sheets[s_name].cell(row=r, column=6, value=f'=VALUE(MID(E{r}, FIND("Total Route =", E{r}) + 14, FIND("M", E{r}) - (FIND("Total Route =", E{r}) + 14)))')
            apply_styling(writer.sheets[s_name], df)
        
        # FAT POLE Recap Logic
        fat_list = []
        pole_list = []
        for sn, df in parsed_dfs.items():
            if "FAT" in sn.upper() and "BOUNDARY" not in sn.upper(): fat_list.extend(df.to_dict('records'))
            elif "POLE" in sn.upper(): pole_list.extend(df.to_dict('records'))
        
        recap = []
        for fat in fat_list:
            found = False
            for pole in pole_list:
                if fat['Latitude'] and pole['Latitude'] and haversine(fat['Latitude'], fat['Longitude'], pole['Latitude'], pole['Longitude']) <= 4.0:
                    recap.append({'FAT Name': fat['Name'], 'POLE Name': pole['Name'], 'Latitude': pole['Latitude'], 'Longitude': pole['Longitude'], 'Description': fat['Description'], 'Total Route (M)': ''})
                    found = True; break
            if not found: recap.append({'FAT Name': fat['Name'], 'POLE Name': '', 'Latitude': '', 'Longitude': '', 'Description': fat['Description'], 'Total Route (M)': ''})
        
        if recap:
            df_fp = pd.DataFrame(recap, columns=FP_HEADERS)
            df_fp.to_excel(writer, sheet_name='FAT & POLE', index=False)
            apply_styling(writer.sheets['FAT & POLE'], df_fp)

    return parsed_dfs